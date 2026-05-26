from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any, Dict, List, Optional, Set, Tuple
from pathlib import Path
import io
import re
import uuid
import random
from datetime import datetime, timezone
import json
import sqlite3
import asyncio
import threading

app = FastAPI()

# Allow all origins so the React dev server and the Cloudflare tunnel can reach the API.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -----------------------------------
# PATHS
# -----------------------------------

BASE_DIR     = Path(__file__).resolve().parent
DATA_DIR     = BASE_DIR / "data"
CLASSES_DIR  = DATA_DIR / "classes"
DATA_DIR.mkdir(exist_ok=True)
CLASSES_DIR.mkdir(exist_ok=True)

REGISTRY_DB  = DATA_DIR / "app.db"
PROJECT_ROOT = Path(__file__).resolve().parents[2]
CTAT_DIR     = PROJECT_ROOT / "frontend_react" / "public" / "CTAT"

# DDL shared by all databases that store step-level interactions.
_STEPS_DDL = """
    CREATE TABLE IF NOT EXISTS steps (
        session_id    TEXT,
        class_code    TEXT,
        problem_id    TEXT,
        step_name     TEXT,
        kc            TEXT,
        hint_per_step INTEGER,
        selection     TEXT,
        action        TEXT,
        input         TEXT,
        correctness   TEXT,
        timestamp     TEXT
    );
"""

# -----------------------------------
# BKT PARAMETERS  (inferred from dataset analysis)
# -----------------------------------

BKT_PARAMS: Dict[str, Dict[str, float]] = {
    "combine_like_terms": {
        "p0": 0.4260, "p_l": 0.2862, "p_f": 0.0697, "p_g": 0.5365, "p_s": 0.0241,
    },
    "expand_eliminate_parentheses": {
        "p0": 0.0611, "p_l": 0.4806, "p_f": 0.0889, "p_g": 0.4311, "p_s": 0.0046,
    },
    "move_constants": {
        "p0": 0.2323, "p_l": 0.0832, "p_f": 0.0149, "p_g": 0.3688, "p_s": 0.0989,
    },
    "normalize_negative_sign": {
        "p0": 0.6464, "p_l": 0.3974, "p_f": 0.1029, "p_g": 0.3190, "p_s": 0.0882,
    },
    "remove_coefficient": {
        "p0": 0.3785, "p_l": 0.0703, "p_f": 0.0095, "p_g": 0.4990, "p_s": 0.0838,
    },
}

# -----------------------------------
# ADAPTIVE EXERCISE SELECTION
# -----------------------------------

STRUGGLE_THRESHOLD   = 0.40   # below → remedial + Easy
MASTERY_THRESHOLD    = 0.80   # avg P(L) threshold to skip a level and advance; difficulty boundary
LEVEL_SKIP_THRESHOLD = 0.85   # stricter per-KC threshold used only for the full-mastery flag
N_ASSIGNED           = 4      # non-remedial exercises per round (N_REGULAR + N_BONUS)
N_REGULAR            = 3      # targeted exercises in regular block
N_BONUS              = 1      # random Medium/Difficult exercise appended to every round
N_REMEDIAL           = 2      # Easy exercises prepended when any KC P(L) < STRUGGLE_THRESHOLD
# Total per round: N_REMEDIAL + N_REGULAR + N_BONUS = 6 when remediation applies, 4 otherwise

# KCs used to evaluate mastery of each level (simplification excluded — no BKT params)
LEVEL_KC_GROUPS: Dict[str, List[str]] = {
    "level1": ["move_constants", "remove_coefficient"],
    "level2": ["combine_like_terms","normalize_negative_sign"],
    "level3": ["expand_eliminate_parentheses"],
}

# Reverse mapping: kc_name → level (derived from LEVEL_KC_GROUPS)
KC_TO_LEVEL: Dict[str, str] = {
    kc: level
    for level, kcs in LEVEL_KC_GROUPS.items()
    for kc in kcs
}

# Both move_constants and remove_coefficient must be sufficiently mastered
# before level-2 KCs (combine_like_terms, normalize_negative_sign) are targeted.
PREREQUISITES: Dict[str, List[str]] = {
    "combine_like_terms":      ["move_constants", "remove_coefficient"],
    "normalize_negative_sign": ["move_constants", "remove_coefficient"],
}
PREREQ_THRESHOLD = 0.50

# Exercises already used in the diagnostic phase — never reassigned
DIAGNOSTIC_EXERCISES = {"level1Difficult_v1", "level2Difficult_v1", "level3Difficult_v1"}

# All exercises that must never appear in personalised rounds
_EXCLUDED_FROM_POOL = DIAGNOSTIC_EXERCISES


def _natural_key(s: str) -> list:
    """Split a string into alternating text/int segments so sort order is 'v2 < v10'."""
    return [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", s)]


def _build_exercise_pool() -> Dict[Tuple[str, str], List[str]]:
    """Scan the CTAT directory and build sorted exercise lists per (level, difficulty)."""
    pool: Dict[Tuple[str, str], List[str]] = {}
    if not CTAT_DIR.exists():
        return pool
    for folder in sorted(CTAT_DIR.iterdir()):
        m = re.match(r"(level\d+)(Easy|Medium|Difficult)$", folder.name)
        if not m:
            continue
        html_dir = folder / "HTML"
        if not html_dir.exists():
            continue
        ids = []
        for f in html_dir.glob("*.html"):
            name = re.sub(r"\.html$", "", f.name)
            name = re.sub(r"^HTML_", "", name)
            if name not in _EXCLUDED_FROM_POOL:
                ids.append(name)
        ids.sort(key=_natural_key)
        pool[(m.group(1), m.group(2))] = ids
    return pool


EXERCISE_POOL: Dict[Tuple[str, str], List[str]] = _build_exercise_pool()

# ---------------------------------------------------------------------------
# KC → exercise-version mapping
#
# Within a (level, difficulty) pool some version ranges target a specific KC.
# Format: (level, difficulty) → { kc_name: (v_min, v_max) inclusive }
# When a KC has no entry here, the full pool is used for that KC.
# ---------------------------------------------------------------------------
KC_VERSION_RANGES: Dict[Tuple[str, str], Dict[str, Tuple[int, int]]] = {
    ("level1", "Easy"):   {"move_constants": (1, 5), "remove_coefficient": (6, 10)},
    ("level1", "Medium"): {"move_constants": (1, 5), "remove_coefficient": (6, 10)},
    # v1–v5 introduce combine_like_terms + normalize_negative_sign;
    # v6–v10 shift to move_constants + remove_coefficient (level1 revision).
    ("level2", "Medium"): {"combine_like_terms": (1, 5), "normalize_negative_sign": (1, 5)},
}


def _version_num(exercise_id: str) -> int:
    m = re.search(r"_v(\d+)$", exercise_id)
    return int(m.group(1)) if m else 0


def _build_kc_exercise_map() -> Dict[Tuple[str, str, str], List[str]]:
    """Build (level, difficulty, kc) → sorted exercise-ID list from KC_VERSION_RANGES."""
    mapping: Dict[Tuple[str, str, str], List[str]] = {}
    for (level, diff), kc_ranges in KC_VERSION_RANGES.items():
        pool = EXERCISE_POOL.get((level, diff), [])
        for kc, (v_min, v_max) in kc_ranges.items():
            mapping[(level, diff, kc)] = [
                ex for ex in pool if v_min <= _version_num(ex) <= v_max
            ]
    return mapping


KC_EXERCISE_MAP: Dict[Tuple[str, str, str], List[str]] = _build_kc_exercise_map()


def _prerequisites_met(kc: str, knowledge_states: Dict[str, float]) -> bool:
    """Return True if all prerequisite KCs for the given KC meet PREREQ_THRESHOLD."""
    return all(
        knowledge_states.get(prereq, 0) >= PREREQ_THRESHOLD
        for prereq in PREREQUISITES.get(kc, [])
    )


def select_exercises(
    knowledge_states: Dict[str, float],
    used_ids: Optional[List[str]] = None,
) -> Tuple[List[str], str, str]:
    """
    Return (problem_ids, level, difficulty) for a student's personalized set.

    The selection has three blocks concatenated:

    REMEDIAL block (N_REMEDIAL = 2 exercises, only when any KC P(L) < STRUGGLE_THRESHOLD):
      Collect Easy exercises for every struggling KC, shuffle, take N_REMEDIAL.

    REGULAR block (N_REGULAR = 3 exercises):
      Determine the student's working level: walk level1 → level2 → level3 and
      skip a level when its average P(L) ≥ MASTERY_THRESHOLD (0.80).  The first
      non-skipped level is the working level.  Difficulty: Easy if avg < STRUGGLE,
      Difficult if avg ≥ MASTERY but level not yet skipped, Medium otherwise.

    BONUS block (N_BONUS = 1 exercise):
      One random exercise from the Medium or Difficult pool (any level), giving the
      student varied exposure beyond their current assigned level.

    Total: N_REMEDIAL + N_REGULAR + N_BONUS = 6 when remediation applies, 4 otherwise.
    Already-completed exercises (used_ids) are excluded throughout.
    """
    excluded = DIAGNOSTIC_EXERCISES | set(used_ids or [])

    # ── Remedial block ────────────────────────────────────────────────────────
    remedial_pool: List[str] = []
    for kc, score in knowledge_states.items():
        if score >= STRUGGLE_THRESHOLD:
            continue
        if not _prerequisites_met(kc, knowledge_states):
            continue  # prerequisites not yet met; regular block will address those first
        kc_level = KC_TO_LEVEL.get(kc)
        if not kc_level:
            continue
        subset = KC_EXERCISE_MAP.get((kc_level, "Easy", kc))
        if subset:
            remedial_pool.extend(ex for ex in subset if ex not in excluded)
        else:
            remedial_pool.extend(
                ex for ex in EXERCISE_POOL.get((kc_level, "Easy"), [])
                if ex not in excluded
            )

    remedial_pool = list(dict.fromkeys(remedial_pool))  # deduplicate, preserve order
    random.shuffle(remedial_pool)
    remedial = remedial_pool[:N_REMEDIAL]

    # Exclude chosen remedial exercises so the regular block never repeats them
    excluded_regular = excluded | set(remedial)

    # ── Regular block ─────────────────────────────────────────────────────────
    chosen_level      = "level3"
    chosen_difficulty = "Difficult"
    regular: List[str] = []

    for level in ("level1", "level2", "level3"):
        kcs = [k for k in LEVEL_KC_GROUPS[level] if k in knowledge_states]
        if not kcs:
            continue

        scores = {kc: knowledge_states[kc] for kc in kcs}
        avg    = sum(scores.values()) / len(scores)

        if avg >= MASTERY_THRESHOLD:
            continue  # level mastered → check next

        if avg < STRUGGLE_THRESHOLD:
            difficulty = "Easy"
        elif avg < MASTERY_THRESHOLD:
            difficulty = "Medium"
        else:
            difficulty = "Difficult"  # avg high but individual KCs not all above skip threshold

        chosen_level      = level
        chosen_difficulty = difficulty

        weak_kcs = [kc for kc in kcs if scores[kc] < MASTERY_THRESHOLD]
        targeted: List[str] = []
        for kc in weak_kcs:
            subset = KC_EXERCISE_MAP.get((level, difficulty, kc))
            if subset:
                targeted.extend(ex for ex in subset if ex not in excluded_regular)
        targeted = list(dict.fromkeys(targeted))

        if not targeted:
            targeted = [
                ex for ex in EXERCISE_POOL.get((level, difficulty), [])
                if ex not in excluded_regular
            ]

        random.shuffle(targeted)
        regular = targeted[:N_REGULAR]
        break
    else:
        # All levels skipped → enrichment with level3Difficult
        pool = [
            ex for ex in EXERCISE_POOL.get(("level3", "Difficult"), [])
            if ex not in excluded_regular
        ]
        random.shuffle(pool)
        regular = pool[:N_REGULAR]

    # ── Bonus block: 1 random Medium or Difficult exercise ───────────────────
    excluded_bonus = excluded | set(remedial) | set(regular)
    bonus_pool: List[str] = []
    for diff in ("Medium", "Difficult"):
        for lvl in ("level1", "level2", "level3"):
            bonus_pool.extend(
                ex for ex in EXERCISE_POOL.get((lvl, diff), [])
                if ex not in excluded_bonus
            )
    bonus_pool = list(dict.fromkeys(bonus_pool))
    random.shuffle(bonus_pool)
    bonus = bonus_pool[:N_BONUS]

    return remedial + regular + bonus, chosen_level, chosen_difficulty


# -----------------------------------
# BKT LOGIC
# -----------------------------------

def bkt_update(p_l: float, correct: bool, params: Dict[str, float]) -> float:
    """
    Perform one Bayesian Knowledge Tracing update step, including a forget term.

    p_l:    current probability the student has learned the KC
    correct: True if the attempt was correct without a hint
    params: dict with p_s (slip), p_g (guess), p_f (forget), p_l (learn)

    Returns the updated P(L) after observing this attempt.
    The Bayes update uses the standard BKT formula:
        P(L|obs) = P(obs|L)*P(L) / P(obs)
    Then the learn/forget transition is applied to give P(L_{t+1}).
    """
    p_s   = params["p_s"]
    p_g   = params["p_g"]
    p_f   = params["p_f"]
    p_lrn = params["p_l"]

    if correct:
        num = p_l * (1.0 - p_s) # probability of knew it and didn't slip 
        den = num + (1.0 - p_l) * p_g # probability of didn't know it but guessed it right. 
    else:
        num = p_l * p_s         # probabilitiy of know it but slipped 
        den = num + (1.0 - p_l) * (1.0 - p_g) # probability of didn't know it and didn't guess right 

    p_post = num / den if den > 0 else p_l
    # Apply learn/forget transition: even a student who knows may forget,
    # and one who does not may spontaneously learn.
    return p_post * (1.0 - p_f) + (1.0 - p_post) * p_lrn # knew it and keep it + didn't know it but learned it


def compute_knowledge_states(session_id: str, db_path: Path) -> Dict[str, float]:
    """
    Run BKT over the full attempt history of a session and return the final P(L) per KC.

    Only ATTEMPT events without a hint (hint_per_step=0) are treated as correct evidence.
    Hint-assisted correct answers do not update knowledge upward because they don't
    distinguish student skill from hint use.
    """
    conn = sqlite3.connect(db_path, timeout=10)
    rows = conn.execute("""
        SELECT kc, correctness, hint_per_step, timestamp
        FROM   steps
        WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
        ORDER  BY timestamp ASC
    """, (session_id,)).fetchall()
    conn.close()

    states = {kc: p["p0"] for kc, p in BKT_PARAMS.items()}
    for kc, correctness, hint_per_step, _ts in rows:
        if kc not in BKT_PARAMS:
            continue
        correct = (correctness == "CORRECT") and (not hint_per_step)
        states[kc] = bkt_update(states[kc], correct, BKT_PARAMS[kc])
    return states


def compute_knowledge_states_with_trace(
    session_id: str, db_path: Path
) -> Tuple[Dict[str, float], List[Dict]]:
    """
    Same as compute_knowledge_states but also returns the full BKT trajectory.

    The trace is one entry per update step with the timestamp, which KC was updated,
    the observation, and the resulting P(L) for all KCs at that point.
    Used for plotting per-student learning curves and research analysis.
    """
    conn = sqlite3.connect(db_path, timeout=10)
    rows = conn.execute("""
        SELECT kc, correctness, hint_per_step, timestamp
        FROM   steps
        WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
        ORDER  BY timestamp ASC
    """, (session_id,)).fetchall()
    conn.close()

    states = {kc: p["p0"] for kc, p in BKT_PARAMS.items()}
    trace: List[Dict] = []
    for i, (kc, correctness, hint_per_step, ts) in enumerate(rows):
        if kc not in BKT_PARAMS:
            continue
        correct = (correctness == "CORRECT") and (not hint_per_step)
        states[kc] = bkt_update(states[kc], correct, BKT_PARAMS[kc])
        trace.append({
            "step_idx":                       i,
            "timestamp":                      ts,
            "kc_updated":                     kc,
            "correctness":                    correctness,
            "hint_per_step":                  hint_per_step,
            "p_move_constants":               round(states.get("move_constants", 0), 4),
            "p_remove_coefficient":           round(states.get("remove_coefficient", 0), 4),
            "p_combine_like_terms":           round(states.get("combine_like_terms", 0), 4),
            "p_normalize_negative_sign":      round(states.get("normalize_negative_sign", 0), 4),
            "p_expand_eliminate_parentheses": round(states.get("expand_eliminate_parentheses", 0), 4),
        })
    return states, trace


# -----------------------------------
# POST-DIAGNOSTIC PROCESSING
# (runs in background after N_FIRST_EXERCISES done)
# -----------------------------------

def process_completed_session(
    session_id: str,
    class_code: str,
    db_path: Path,
    used_ids: Optional[List[str]] = None,
    report_phase: int = 1,
) -> None:
    """
    Run BKT, select the next exercise set, store the assignment, and update the report.

    Called in a background thread after each completed exercise round:
      - round 0 (diagnostics done)    → report_phase=1, used_ids=[]
      - round 1+ (personalized done)  → report_phase=2, used_ids=all previously done IDs

    The assignment row is written (or overwritten) here so the WaitingPage gets
    `ready: true` as soon as BKT finishes.
    """
    round_label = "diagnostics" if report_phase == 1 else f"personalized round {report_phase - 1}"
    print(f"[BKT] Processing after {round_label}  session={session_id[:8]}  class={class_code}")

    # 1. BKT over all attempts so far (with full trajectory for research logging)
    knowledge_states, bkt_trace = compute_knowledge_states_with_trace(session_id, db_path)
    print(f"[BKT] States: { {k: round(v,3) for k,v in knowledge_states.items()} }")

    # 2. Select next personalised exercise set (excluding already-used IDs)
    problem_ids, level, difficulty = select_exercises(knowledge_states, used_ids)

    all_mastered = all(
        knowledge_states.get(kc, 0) >= LEVEL_SKIP_THRESHOLD for kc in BKT_PARAMS
    )
    if all_mastered:
        print(f"[BKT] MASTERY ACHIEVED  session={session_id[:8]}")
    print(f"[BKT] Assignment → {level} {difficulty}: {problem_ids}")

    # 3. Store assignment (with mastery flag) and BKT trace
    conn = sqlite3.connect(db_path, timeout=10)
    # Ensure bkt_trace table exists for DBs created before this schema version
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bkt_trace (
            session_id                     TEXT,
            phase                          INTEGER,
            step_idx                       INTEGER,
            timestamp                      TEXT,
            kc_updated                     TEXT,
            correctness                    TEXT,
            hint_per_step                  INTEGER,
            p_move_constants               REAL,
            p_remove_coefficient           REAL,
            p_combine_like_terms           REAL,
            p_normalize_negative_sign      REAL,
            p_expand_eliminate_parentheses REAL
        )
    """)
    conn.execute("""
        INSERT OR REPLACE INTO assignments
            (session_id, class_code, level, difficulty, problem_ids, assigned_at, mastery)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        session_id, class_code, level, difficulty,
        json.dumps(problem_ids),
        datetime.now(timezone.utc).isoformat(),
        1 if all_mastered else 0,
    ))
    for entry in bkt_trace:
        conn.execute("""
            INSERT INTO bkt_trace
                (session_id, phase, step_idx, timestamp, kc_updated,
                 correctness, hint_per_step,
                 p_move_constants, p_remove_coefficient, p_combine_like_terms,
                 p_normalize_negative_sign, p_expand_eliminate_parentheses)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session_id, report_phase, entry["step_idx"], entry["timestamp"],
            entry["kc_updated"], entry["correctness"], entry["hint_per_step"],
            entry["p_move_constants"], entry["p_remove_coefficient"],
            entry["p_combine_like_terms"], entry["p_normalize_negative_sign"],
            entry["p_expand_eliminate_parentheses"],
        ))
    conn.commit()
    conn.close()

    # 4. Write / update the per-student Excel report
    try:
        generate_student_report(session_id, class_code, db_path, phase=report_phase)
    except Exception as exc:
        print(f"[REPORT] Error generating student report: {exc}")


# -----------------------------------
# REPORT GENERATION
# -----------------------------------

def generate_student_report(
    session_id: str, class_code: str, db_path: Path, phase: int
) -> Optional[Path]:
    """
    Create (phase=1) or update (phase=2) a per-student Excel workbook.

    Sheet "Resum"  — one summary row per phase (diagnostic / personalized).
    Sheet "Intents" — every attempt event recorded for this student.

    File name: student_{safe_student_id}_{class_code}.xlsx
    Phase 1 is called right after BKT runs on the 3 diagnostic exercises.
    Phase 2 is called when the last personalized exercise is completed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[REPORT] openpyxl not installed — run: pip install openpyxl")
        raise

    conn = sqlite3.connect(db_path, timeout=10)

    row = conn.execute(
        "SELECT student_id, created_at, completed_problems FROM sessions WHERE session_id = ?",
        (session_id,),
    ).fetchone()
    if not row:
        conn.close()
        return None
    student_id, created_at, completed_problems_json = row
    completions = json.loads(completed_problems_json)  # list of {"problem_id", "completed_at"}

    assignment = conn.execute(
        "SELECT level, difficulty FROM assignments WHERE session_id = ?",
        (session_id,),
    ).fetchone()

    stats = conn.execute("""
        SELECT
            COUNT(*),
            COUNT(CASE WHEN correctness = 'CORRECT' AND hint_per_step = 0 THEN 1 END),
            COUNT(CASE WHEN hint_per_step = 1 THEN 1 END)
        FROM steps WHERE session_id = ?
    """, (session_id,)).fetchone()
    total_attempts, correct_no_hint, total_hints = stats
    accuracy = round(correct_no_hint / total_attempts, 3) if total_attempts else 0.0

    last_completed_at = completions[-1]["completed_at"] if completions else None
    try:
        t0 = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        t1 = datetime.fromisoformat(last_completed_at.replace("Z", "+00:00")) if last_completed_at else t0
        duration_min = round((t1 - t0).total_seconds() / 60, 2)
    except Exception:
        duration_min = None

    steps_rows = conn.execute("""
        SELECT session_id, class_code, problem_id, step_name, kc,
               hint_per_step, selection, action, input, correctness, timestamp
        FROM   steps
        WHERE  session_id = ?
        ORDER  BY timestamp ASC
    """, (session_id,)).fetchall()

    conn.close()

    knowledge_states = compute_knowledge_states(session_id, db_path)

    kc_order = list(BKT_PARAMS.keys())
    resum_headers = [
        "Fase", "Alumne", "Hora finalització", "Durada (min)",
        "Exercicis fets", "Total intents", "Total pistes", "Precisió",
    ] + [f"P(L) {kc}" for kc in kc_order] + ["Nivell assignat", "Dificultat assignada"]

    summary_data: Dict[str, Any] = {
        "Fase":               "Diagnosi" if phase == 1 else f"Personalitzada {phase - 1}",
        "Alumne":             student_id,
        "Hora finalització":  last_completed_at or "",
        "Durada (min)":       duration_min,
        "Exercicis fets":     len(completions),
        "Total intents":      total_attempts,
        "Total pistes":       total_hints,
        "Precisió":           accuracy,
        **{f"P(L) {kc}": round(v, 4) for kc, v in knowledge_states.items()},
        "Nivell assignat":       assignment[0] if assignment else None,
        "Dificultat assignada":  assignment[1] if assignment else None,
    }

    intents_headers = [
        "session_id", "class_code", "problem_id", "step_name", "kc",
        "hint_per_step", "selection", "action", "input", "correctness", "timestamp",
    ]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    blue   = PatternFill("solid", fgColor="1565C0")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    green  = PatternFill("solid", fgColor="E8F5E9")
    w_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Border(**{s: Side(style="thin") for s in ("left", "right", "top", "bottom")})

    safe_name   = re.sub(r"[^\w\-]", "_", student_id)
    # db_path is the main class DB; Excel goes in the students/ subfolder beside it
    report_path = db_path.parent / "students" / f"student_{safe_name}_{class_code}.xlsx"

    def _style_data_cell(cell, header: str, row_idx: int) -> None:
        cell.border    = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if header.startswith("P(L)"):
            cell.fill = yellow
        elif "assignat" in header.lower() or "assignada" in header.lower():
            cell.fill = green
        else:
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 == 0 else "F5F5F5")

    if phase == 1 or not report_path.exists():
        wb = openpyxl.Workbook()

        # ── Sheet "Resum" ──────────────────────────────────────────────────
        ws_resum = wb.active
        ws_resum.title = "Resum"

        for col, h in enumerate(resum_headers, 1):
            c = ws_resum.cell(row=1, column=col, value=h)
            c.fill = blue; c.font = w_font; c.alignment = center; c.border = thin
        ws_resum.row_dimensions[1].height = 36

        for col, h in enumerate(resum_headers, 1):
            c = ws_resum.cell(row=2, column=col, value=summary_data.get(h))
            _style_data_cell(c, h, 2)

        # ── Sheet "Intents" ────────────────────────────────────────────────
        ws_int = wb.create_sheet("Intents")

        for col, h in enumerate(intents_headers, 1):
            c = ws_int.cell(row=1, column=col, value=h)
            c.fill = blue; c.font = w_font; c.alignment = center; c.border = thin

        for ri, attempt in enumerate(steps_rows, 2):
            for col, val in enumerate(attempt, 1):
                c = ws_int.cell(row=ri, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center")

        for ws in (ws_resum, ws_int):
            for col_cells in ws.columns:
                width = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = max(14, width + 2)

    else:
        # Phase 2: load existing workbook and append new data
        wb       = openpyxl.load_workbook(report_path)
        ws_resum = wb["Resum"]
        ws_int   = wb["Intents"]

        # Add phase-2 summary row
        next_row = ws_resum.max_row + 1
        for col, h in enumerate(resum_headers, 1):
            c = ws_resum.cell(row=next_row, column=col, value=summary_data.get(h))
            _style_data_cell(c, h, next_row)

        # Append only attempt rows not yet written (by offset)
        existing_attempts = ws_int.max_row - 1  # subtract header row
        new_rows = steps_rows[existing_attempts:]
        for attempt in new_rows:
            ri = ws_int.max_row + 1
            for col, val in enumerate(attempt, 1):
                c = ws_int.cell(row=ri, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(report_path)
    print(f"[REPORT] Student report phase {phase} → {report_path.name}")
    return report_path


def generate_class_report(class_code: str, db_path: Path) -> Optional[Path]:
    """
    Build a colour-coded Excel workbook with one row per student and save it to DATA_DIR.

    Columns: student ID, completion time, duration, attempt counts, accuracy,
    assigned level/difficulty, and the final P(L) for every KC.
    The file name is timestamped so multiple reports coexist without overwriting.
    Returns None if no students have completed the diagnostic yet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[REPORT] openpyxl not installed — run: pip install openpyxl")
        raise

    conn = sqlite3.connect(db_path, timeout=10)
    sessions = conn.execute(
        "SELECT session_id, student_id, created_at FROM sessions WHERE class_code = ?",
        (class_code,),
    ).fetchall()

    report_rows = []
    for session_id, student_id, created_at in sessions:
        cp_row = conn.execute(
            "SELECT completed_problems FROM sessions WHERE session_id = ?",
            (session_id,),
        ).fetchone()
        completions = json.loads(cp_row[0]) if cp_row else []
        if not completions:
            continue

        last_ts = completions[-1]["completed_at"]
        try:
            t0 = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            t1 = datetime.fromisoformat(last_ts.replace("Z", "+00:00"))
            duration_min = round((t1 - t0).total_seconds() / 60, 2)
        except Exception:
            duration_min = None

        stats = conn.execute("""
            SELECT
                COUNT(*),
                COUNT(CASE WHEN correctness = 'CORRECT' AND hint_per_step = 0 THEN 1 END),
                COUNT(CASE WHEN hint_per_step = 1 THEN 1 END)
            FROM steps WHERE session_id = ?
        """, (session_id,)).fetchone()
        total_attempts, correct_no_hint, total_hints = stats
        accuracy = round(correct_no_hint / total_attempts, 3) if total_attempts else 0.0

        assignment = conn.execute(
            "SELECT level, difficulty, problem_ids FROM assignments WHERE session_id = ?",
            (session_id,),
        ).fetchone()
        assigned_level = assignment[0] if assignment else None
        assigned_diff  = assignment[1] if assignment else None

        states = compute_knowledge_states(session_id, db_path)

        report_rows.append({
            "student_id":      student_id,
            "completion_time": last_ts,
            "duration_min":    duration_min,
            "total_attempts":  total_attempts,
            "total_hints":     total_hints,
            "accuracy":        accuracy,
            "assigned_level":  assigned_level,
            "assigned_difficulty": assigned_diff,
            **{f"p_{kc}": round(v, 4) for kc, v in states.items()},
        })

    conn.close()

    if not report_rows:
        return None

    kc_cols = [f"p_{kc}" for kc in BKT_PARAMS]
    headers = [
        "student_id", "completion_time", "duration_min",
        "total_attempts", "total_hints", "accuracy",
        "assigned_level", "assigned_difficulty",
    ] + kc_cols

    LABELS = {
        "student_id":           "Student",
        "completion_time":      "Completion Time",
        "duration_min":         "Duration (min)",
        "total_attempts":       "Total Attempts",
        "total_hints":          "Total Hints",
        "accuracy":             "Accuracy (no hint)",
        "assigned_level":       "Assigned Level",
        "assigned_difficulty":  "Assigned Difficulty",
        **{f"p_{kc}": f"P(L) {kc}" for kc in BKT_PARAMS},
    }

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Knowledge States"

    blue   = PatternFill("solid", fgColor="1565C0")
    yellow = PatternFill("solid", fgColor="FFF9C4")
    green  = PatternFill("solid", fgColor="E8F5E9")
    w_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(**{s: Side(style="thin") for s in ("left","right","top","bottom")})

    for col, key in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=LABELS.get(key, key))
        c.fill = blue; c.font = w_font; c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 36

    for ri, row in enumerate(report_rows, 2):
        bg = PatternFill("solid", fgColor="FFFFFF" if ri % 2 == 0 else "F5F5F5")
        for col, key in enumerate(headers, 1):
            c = ws.cell(row=ri, column=col, value=row.get(key))
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = yellow if key.startswith("p_") else (green if key.startswith("assigned") else bg)

    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = max(14, width + 2)

    ts_str = datetime.now().strftime("%Hh%M_%d-%m-%Y")
    # db_path is the main class DB; class report goes in the same folder
    report_path = db_path.parent / f"report_{ts_str}_{class_code}.xlsx"
    wb.save(report_path)
    print(f"[REPORT] Saved → {report_path.name}")
    return report_path


# -----------------------------------
# DATABASE HELPERS
# -----------------------------------

def init_registry():
    """
    Create the global registry database that maps each session_id to its class DB file.

    This is a single lightweight DB (app.db) that survives server restarts and lets
    any endpoint look up which per-class SQLite file to open for a given session_id.
    """
    conn = sqlite3.connect(REGISTRY_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS session_registry (
            session_id TEXT PRIMARY KEY,
            class_code TEXT,
            db_file    TEXT
        )
    """)
    conn.commit()
    conn.close()
    print(f"Registry DB ready: {REGISTRY_DB}")


def _init_steps_db(path: Path) -> None:
    """Create a lightweight DB that only holds the steps table."""
    conn = sqlite3.connect(path, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.executescript(_STEPS_DDL)
    conn.commit()
    conn.close()


def create_class_db(class_code: str) -> Path:
    """
    Build the full folder structure for one class session and return the main DB path.

    Layout inside CLASSES_DIR:
        {class_code}_{HHhMM_DD-MM-YYYY}/
            {class_code}_{HHhMM_DD-MM-YYYY}.db   ← main class DB (includes first_analysis table)
            students/                             ← one DB per student, created on join
    """
    ts = datetime.now()
    folder_name = f"{class_code}_{ts.strftime('%Hh%M_%d-%m-%Y')}"
    class_folder = CLASSES_DIR / folder_name
    class_folder.mkdir(parents=True, exist_ok=True)
    (class_folder / "students").mkdir(exist_ok=True)

    db_path = class_folder / f"{folder_name}.db"

    # Main class DB — sessions + steps + assignments
    conn = sqlite3.connect(db_path, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sessions (
            session_id         TEXT PRIMARY KEY,
            class_code         TEXT,
            student_id         TEXT,
            created_at         TEXT,
            hand_raised        INTEGER NOT NULL DEFAULT 0,
            completed_problems TEXT    NOT NULL DEFAULT '[]'
        );
        CREATE TABLE IF NOT EXISTS assignments (
            session_id   TEXT PRIMARY KEY,
            class_code   TEXT,
            level        TEXT,
            difficulty   TEXT,
            problem_ids  TEXT,
            assigned_at  TEXT,
            mastery      INTEGER NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS bkt_trace (
            session_id                     TEXT,
            phase                          INTEGER,
            step_idx                       INTEGER,
            timestamp                      TEXT,
            kc_updated                     TEXT,
            correctness                    TEXT,
            hint_per_step                  INTEGER,
            p_move_constants               REAL,
            p_remove_coefficient           REAL,
            p_combine_like_terms           REAL,
            p_normalize_negative_sign      REAL,
            p_expand_eliminate_parentheses REAL
        );
        CREATE TABLE IF NOT EXISTS first_analysis (
            session_id    TEXT,
            class_code    TEXT,
            problem_id    TEXT,
            step_name     TEXT,
            kc            TEXT,
            hint_per_step INTEGER,
            selection     TEXT,
            action        TEXT,
            input         TEXT,
            correctness   TEXT,
            timestamp     TEXT
        );
    """ + _STEPS_DDL)
    conn.commit()
    conn.close()

    print(f"[CLASS DB] folder={folder_name}")
    return db_path


def get_db_for_session(session_id: str) -> Optional[Path]:
    """Return the Path of the class DB for a session, checking in-memory first then the registry."""
    if session_id in SESSIONS:
        return Path(SESSIONS[session_id]["db_path"])
    conn = sqlite3.connect(REGISTRY_DB)
    row = conn.execute(
        "SELECT db_file FROM session_registry WHERE session_id = ?", (session_id,)
    ).fetchone()
    conn.close()
    return DATA_DIR / row[0] if row else None


def get_class_code_for_session(session_id: str) -> Optional[str]:
    """Return the class_code for a session, checking in-memory first then the registry."""
    if session_id in SESSIONS:
        return SESSIONS[session_id].get("class_code")
    conn = sqlite3.connect(REGISTRY_DB)
    row = conn.execute(
        "SELECT class_code FROM session_registry WHERE session_id = ?", (session_id,)
    ).fetchone()
    conn.close()
    return row[0] if row else None


init_registry()

# -----------------------------------
# IN-MEMORY STATE
# -----------------------------------

ACTIVE_CLASSES: Dict[str, Path] = {}
# session_id → {class_code, student_id, db_path, hints_seen}
SESSIONS: Dict[str, Dict[str, Any]] = {}
CLASS_STREAMS: Dict[str, List[asyncio.Queue]] = {}
CLASS_ASSIGNMENTS: Dict[str, str] = {}
ENDED_CLASSES: Set[str] = set()  # class_codes for which the teacher has ended the session
CLASS_MESSAGES_ENABLED: Dict[str, bool] = {}  # class_code → messages enabled flag (A/B test)

N_FIRST_EXERCISES = 3

LEVEL_CLASS_PROBLEM: Dict[int, str] = {
    1: "level1Difficult_v1",
    2: "level2Difficult_v1",
    3: "level3Difficult_v1",
}

# -----------------------------------
# REQUEST MODELS
# -----------------------------------

class StartSessionRequest(BaseModel):
    class_code: str
    student_id: str

class DiagnosticCompleteRequest(BaseModel):
    session_id: str

class LogsRequest(BaseModel):
    session_id: str
    events: List[Dict[str, Any]]

class AssignRequest(BaseModel):
    level: int


# -----------------------------------
# ROUTES
# -----------------------------------

@app.get("/api/health")
def health():
    """Simple liveness check — the frontend calls this to verify the server is reachable."""
    return {"status": "ok"}


@app.post("/api/classroom/{class_code}/start")
def start_class(class_code: str):
    """
    Teacher endpoint: initialise or resume a class session.

    If a folder for this class_code already exists from TODAY, that session is
    resumed (its DB is reloaded into ACTIVE_CLASSES) so students can reconnect
    and continue from where they left off.  If no today-folder exists a fresh DB
    is created as before.

    "Today" is matched by the DD-MM-YYYY suffix in the folder name so a teacher
    who uses the same class_code on a different day always gets a clean session.
    """
    today_str = datetime.now().strftime("%d-%m-%Y")
    # Look for any folder created today for this class_code
    existing_today = sorted(
        [
            f for f in CLASSES_DIR.iterdir()
            if f.is_dir()
            and f.name.startswith(f"{class_code}_")
            and today_str in f.name
        ],
        key=lambda p: p.name,
    )

    if existing_today:
        latest_folder = existing_today[-1]
        # The DB file has the same stem as the folder
        db_candidates = list(latest_folder.glob("*.db"))
        if db_candidates:
            db_path = db_candidates[0]
            ACTIVE_CLASSES[class_code] = db_path
            CLASS_MESSAGES_ENABLED.setdefault(class_code, True)
            print(f"[CLASS RESUMED] {class_code}  →  {db_path.name}")
            return {"status": "resumed", "class_code": class_code, "db": db_path.name}

    # No existing session today → create fresh
    db_path = create_class_db(class_code)
    ACTIVE_CLASSES[class_code] = db_path
    CLASS_MESSAGES_ENABLED[class_code] = True
    print(f"[CLASS STARTED] {class_code}  →  {db_path.name}")
    return {"status": "started", "class_code": class_code, "db": db_path.name}


@app.post("/api/session/start")
def start_session(req: StartSessionRequest):
    """
    Student endpoint: create or resume a session and return the exercise queue.

    Requires the teacher to have already started (or resumed) the class.
    Returns 403 otherwise.

    RESUME path (network-error recovery):
      If a session for this student_id already exists in the active class DB the
      existing session is restored into in-memory state and returned together with:
        - completed_problems  list of already-finished exercises
        - assignment          current personalised set (if BKT already ran), or null
        - resumed             true

      The frontend uses these fields to skip exercises the student already did and
      route them directly to the correct page (tutor / waiting).

      Edge case: if the diagnostics are done but the assignment is missing (server
      crashed while BKT was running in the background) the BKT computation is
      automatically re-triggered so the WaitingPage will receive the result.

    NEW SESSION path (normal flow):
      Creates a fresh session row, registers it in the registry DB, and returns
      the three diagnostic problem IDs.
    """
    if req.class_code not in ACTIVE_CLASSES:
        raise HTTPException(
            status_code=403,
            detail="Aquesta classe no ha iniciat sessió. Espera que el professor l'iniciï.",
        )

    db_path      = ACTIVE_CLASSES[req.class_code]
    class_folder = db_path.parent
    safe_sid     = re.sub(r"[^\w\-]", "_", req.student_id)

    # ── Check for an existing session for this student in this class ──────────
    conn = sqlite3.connect(db_path, timeout=10)
    existing = conn.execute(
        "SELECT session_id, completed_problems FROM sessions WHERE class_code = ? AND student_id = ?",
        (req.class_code, req.student_id),
    ).fetchone()

    if existing:
        session_id, completed_json = existing
        completed = json.loads(completed_json or "[]")

        # Fetch current assignment (may be None if BKT hasn't run yet or round just ended)
        assignment_row = conn.execute(
            "SELECT level, difficulty, problem_ids, mastery FROM assignments WHERE session_id = ?",
            (session_id,),
        ).fetchone()
        conn.close()

        assignment = None
        if assignment_row:
            assignment = {
                "level":       assignment_row[0],
                "difficulty":  assignment_row[1],
                "problem_ids": json.loads(assignment_row[2]),
                "mastery":     bool(assignment_row[3]),
            }

        # Restore session into in-memory SESSIONS cache
        student_db_path = class_folder / "students" / f"{safe_sid}.db"
        SESSIONS[session_id] = {
            "class_code":      req.class_code,
            "student_id":      req.student_id,
            "db_path":         str(db_path),
            "student_db_path": str(student_db_path),
            "hints_seen":      {},  # ephemeral — cannot be recovered; safe to reset
        }

        completed_ids    = [c["problem_id"] for c in completed]
        n_done           = len(completed)
        diagnostics_done = n_done >= N_FIRST_EXERCISES

        # Re-trigger BKT if diagnostics finished but assignment is missing
        # (server crashed while BKT was running in the background thread)
        if diagnostics_done and assignment is None:
            personalized_done = [pid for pid in completed_ids if pid not in DIAGNOSTIC_EXERCISES]
            rounds_done   = len(personalized_done) // N_ASSIGNED if personalized_done else 0
            report_phase  = 1 + rounds_done
            print(
                f"[RESUME] Re-triggering BKT  session={session_id[:8]}"
                f"  phase={report_phase}  used={personalized_done}"
            )
            threading.Thread(
                target=process_completed_session,
                args=(session_id, req.class_code, db_path, personalized_done, report_phase),
                daemon=True,
            ).start()

        print(
            f"[SESSION RESUMED] class={req.class_code}  student={req.student_id}"
            f"  session={session_id[:8]}  done={n_done}"
        )
        return {
            "session_id":        session_id,
            "problem_ids":       list(DIAGNOSTIC_EXERCISES),  # kept for API compat
            "completed_problems": completed,
            "assignment":        assignment,
            "resumed":           True,
        }

    conn.close()

    # ── New session ───────────────────────────────────────────────────────────
    session_id      = str(uuid.uuid4())
    student_db_path = class_folder / "students" / f"{safe_sid}.db"
    _init_steps_db(student_db_path)

    SESSIONS[session_id] = {
        "class_code":      req.class_code,
        "student_id":      req.student_id,
        "db_path":         str(db_path),
        "student_db_path": str(student_db_path),
        "hints_seen":      {},
    }

    conn = sqlite3.connect(db_path, timeout=10)
    conn.execute(
        "INSERT INTO sessions (session_id, class_code, student_id, created_at) VALUES (?, ?, ?, ?)",
        (session_id, req.class_code, req.student_id, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    conn.close()

    rel_path = str(db_path.relative_to(DATA_DIR))
    conn = sqlite3.connect(REGISTRY_DB)
    conn.execute(
        "INSERT OR REPLACE INTO session_registry (session_id, class_code, db_file) VALUES (?, ?, ?)",
        (session_id, req.class_code, rel_path),
    )
    conn.commit()
    conn.close()

    print(f"[SESSION NEW] class={req.class_code}  student={req.student_id}  folder={class_folder.name}")
    return {
        "session_id":        session_id,
        "problem_ids":       ["level1Difficult_v1", "level2Difficult_v1", "level3Difficult_v1"],
        "completed_problems": [],
        "assignment":        None,
        "resumed":           False,
    }


@app.post("/api/logs")
async def logs(req: LogsRequest):
    """
    Receive and persist CTAT interaction events from the student's browser.

    Handles two event kinds:
      - CTAT_PROBLEM_DONE: records a completion row; triggers BKT + assignment
        once N_FIRST_EXERCISES completions are reached.
      - CTAT_LOG_EVENT: parses the XML payload to extract event_type; only ATTEMPT
        events are stored in the steps table with KC, correctness, and hint_per_step.

    hint_per_step is set to 1 for any step where the student requested a hint
    before submitting, so BKT can exclude those from the learning signal.
    Processing runs in a background thread to avoid blocking the SSE loop.
    """
    db_path = get_db_for_session(req.session_id)
    if db_path is None or not db_path.exists():
        raise HTTPException(status_code=404, detail="Session not found.")

    class_code   = get_class_code_for_session(req.session_id)
    session_meta = SESSIONS.get(req.session_id, {})
    hints_seen: Dict[str, set] = session_meta.get("hints_seen", {})

    student_db_path = Path(session_meta["student_db_path"]) if session_meta.get("student_db_path") else None

    conn        = sqlite3.connect(db_path, timeout=10)
    cursor      = conn.cursor()
    student_conn = sqlite3.connect(student_db_path, timeout=10) if student_db_path else None

    print(f"\n[LOGS] {len(req.events)} event(s)  session={req.session_id[:8]}  class={class_code}")

    sorted_events  = sorted(req.events, key=lambda e: e.get("ts", ""))
    trigger_report = False
    # Tracks hint_per_step from the most recent ATTEMPT so it can be inherited
    # by the paired RESULT event (which carries the KC but not the hint flag).
    step_hints: Dict[str, int] = {}

    for e in sorted_events:
        kind = e.get("kind")
        p    = e.get("payload", {})

        # ── Problem done ────────────────────────────────────────────────────
        if kind == "CTAT_PROBLEM_DONE":
            problem_id = p.get("problemId")
            print(f"  [DONE]  session={req.session_id[:8]}  problem={problem_id}")
            row = cursor.execute(
                "SELECT completed_problems FROM sessions WHERE session_id = ?",
                (req.session_id,),
            ).fetchone()
            completed = json.loads(row[0]) if row else []
            completed.append({"problem_id": problem_id, "completed_at": e.get("ts")})
            cursor.execute(
                "UPDATE sessions SET completed_problems = ?, hand_raised = 0 WHERE session_id = ?",
                (json.dumps(completed), req.session_id),
            )
            trigger_report = True
            continue

        if kind != "CTAT_LOG_EVENT":
            continue

        xml = p.get("xml", "")

        m = re.search(r'name="([^"]+)"', xml)
        event_type = m.group(1) if m else None

        kc_m = re.search(
            r'<skill[^>]*>.*?<name[^>]*>\s*(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?\s*</name>',
            xml, re.DOTALL,
        )
        if not kc_m:
            kc_m = re.search(
                r'<skill_label[^>]*>\s*(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?\s*</skill_label>',
                xml, re.DOTALL,
            )
        # CTAT rule format is "model_name kc_name" (e.g. "move_constants move_constants").
        # Take the last token so the value matches the BKT_PARAMS keys.
        kc_raw = kc_m.group(1).strip() if kc_m else (p.get("kc") or "")
        kc_parts = kc_raw.strip().split()
        kc = kc_parts[-1] if kc_parts else None

        problem_id = p.get("problemId")
        selection  = p.get("selection")
        hint_per_step = None

        correctness = None
        if "INCORRECT" in xml:
            correctness = "INCORRECT"
        elif "CORRECT" in xml:
            correctness = "CORRECT"

        # ── Hint request ────────────────────────────────────────────────────
        if event_type == "HINT_REQUEST":
            if problem_id and selection:
                hints_seen.setdefault(problem_id, set()).add(selection)
            print(
                f"  [HINT_REQUEST]  session={req.session_id[:8]}"
                f"  problem={problem_id}  sel={selection}  kc={kc}"
            )

        # ── Student attempt — track hint state only; DB insert deferred to RESULT ─
        elif event_type == "ATTEMPT":
            hint_per_step = 1 if selection in hints_seen.get(problem_id, set()) else 0
            if problem_id is not None:
                step_hints[problem_id] = hint_per_step
            print(
                f"  [ATTEMPT]  session={req.session_id[:8]}"
                f"  problem={problem_id}  sel={selection}"
                f"  kc={kc}  hint={bool(hint_per_step)}  input={p.get('input')!r}"
                f"  → {correctness or 'n/a'}"
            )
            # kc and correctness are None on ATTEMPT events — they only arrive on the
            # paired RESULT event below, so we insert there (not here).

        # ── Tutor evaluation (RESULT / TUTOR_MSG) — kc and correctness known; store now
        elif kc:
            hint_per_step = step_hints.pop(problem_id, 0) if problem_id else 0
            print(
                f"  [{event_type}]  session={req.session_id[:8]}"
                f"  problem={problem_id}  kc={kc}"
                f"  → {correctness or 'n/a'}  hint={bool(hint_per_step)}"
            )
            _step_sql = """
                INSERT INTO steps
                    (session_id, class_code, problem_id, step_name,
                     kc, hint_per_step, selection, action, input, correctness, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            _step_row = (
                req.session_id, class_code, problem_id,
                selection, kc, hint_per_step,
                selection, p.get("action"), p.get("input"), correctness, e.get("ts"),
            )
            # Main class DB
            cursor.execute(_step_sql, _step_row)
            # Per-student DB
            if student_conn:
                student_conn.execute(_step_sql, _step_row)
            # First analysis table (diagnostic exercises only)
            if problem_id in DIAGNOSTIC_EXERCISES:
                cursor.execute("""
                    INSERT INTO first_analysis
                        (session_id, class_code, problem_id, step_name,
                         kc, hint_per_step, selection, action, input, correctness, timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, _step_row)

    conn.commit()
    if student_conn:
        student_conn.commit()

    # ── Trigger BKT + new assignment after every completed exercise round ────
    if trigger_report:
        row = cursor.execute(
            "SELECT completed_problems FROM sessions WHERE session_id = ?",
            (req.session_id,),
        ).fetchone()
        completed = json.loads(row[0]) if row else []
        n_done = len(completed)
        all_completed_ids = [c["problem_id"] for c in completed]

        if n_done == N_FIRST_EXERCISES:
            # Diagnostic round done → first BKT run, phase-1 report
            conn.close()
            if student_conn: student_conn.close()
            loop = asyncio.get_running_loop()
            loop.run_in_executor(
                None, process_completed_session,
                req.session_id, class_code, db_path, [], 1,
            )
            return {"status": "saved", "assignment_triggered": True}

        elif n_done > N_FIRST_EXERCISES:
            personalized_done = [pid for pid in all_completed_ids if pid not in DIAGNOSTIC_EXERCISES]

            assignment_row = cursor.execute(
                "SELECT problem_ids FROM assignments WHERE session_id = ?",
                (req.session_id,),
            ).fetchone()

            if assignment_row:
                current_assigned = json.loads(assignment_row[0])

                if set(current_assigned).issubset(set(all_completed_ids)):
                    # Current personalized round finished → clear assignment so
                    # WaitingPage shows the spinner while BKT runs in background.
                    cursor.execute(
                        "DELETE FROM assignments WHERE session_id = ?",
                        (req.session_id,),
                    )
                    conn.commit()
                    conn.close()
                    if student_conn: student_conn.close()

                    # report_phase 2+ means "append row to existing file"
                    rounds_done = len(personalized_done) // N_ASSIGNED
                    report_phase = 1 + rounds_done  # phase 2, 3, 4 …

                    loop = asyncio.get_running_loop()
                    loop.run_in_executor(
                        None, process_completed_session,
                        req.session_id, class_code, db_path,
                        personalized_done, report_phase,
                    )
                    return {"status": "saved", "next_round_triggered": True}

    conn.close()
    if student_conn: student_conn.close()
    return {"status": "saved"}


@app.get("/api/session/{session_id}/assignment")
def get_assignment(session_id: str):
    """Student polls this after the waiting page to get their personalised exercise list."""
    db_path = get_db_for_session(session_id)
    if db_path is None or not db_path.exists():
        raise HTTPException(status_code=404, detail="Session not found.")

    conn = sqlite3.connect(db_path, timeout=10)
    row  = conn.execute(
        "SELECT level, difficulty, problem_ids, mastery FROM assignments WHERE session_id = ?",
        (session_id,),
    ).fetchone()
    conn.close()

    if row is None:
        return {"ready": False}

    return {
        "ready":       True,
        "level":       row[0],
        "difficulty":  row[1],
        "problem_ids": json.loads(row[2]),
        "mastery":     bool(row[3]),
    }


@app.post("/api/session/{session_id}/raise-hand")
def raise_hand(session_id: str):
    """Student signals they need help after 3 failed attempts on the same step."""
    db_path = get_db_for_session(session_id)
    if db_path is None or not db_path.exists():
        raise HTTPException(status_code=404, detail="Session not found.")
    conn = sqlite3.connect(db_path, timeout=10)
    try:
        conn.execute(
            "UPDATE sessions SET hand_raised = 1 WHERE session_id = ?", (session_id,)
        )
        conn.commit()
    except Exception:
        pass
    conn.close()
    print(f"[HAND] Raised: session={session_id[:8]}")
    return {"status": "raised"}


@app.post("/api/session/diagnostic-complete")
def diagnostic_complete(req: DiagnosticCompleteRequest):
    return {"assigned_problem_ids": []}


# -----------------------------------
# CLASSROOM ENDPOINTS
# -----------------------------------

@app.post("/api/classroom/{class_code}/assign")
async def assign_problem(class_code: str, req: AssignRequest):
    """
    Teacher endpoint: push a shared exercise to all connected students via SSE.

    Stores the assignment in CLASS_ASSIGNMENTS so students who connect later
    receive it immediately without the teacher needing to re-broadcast.
    """
    problem_id = LEVEL_CLASS_PROBLEM.get(req.level)
    if not problem_id:
        raise HTTPException(status_code=400, detail="Invalid level. Must be 1, 2, or 3.")
    CLASS_ASSIGNMENTS[class_code] = problem_id
    for q in CLASS_STREAMS.get(class_code, []):
        await q.put({"type": "problem_assigned", "problem_id": problem_id})
    return {"problem_id": problem_id, "students_notified": len(CLASS_STREAMS.get(class_code, []))}


@app.get("/api/classroom/{class_code}/stream")
async def classroom_stream(class_code: str, request: Request):
    """
    Student endpoint: open a Server-Sent Events connection to receive real-time assignments.

    Each student gets an independent asyncio.Queue. A 25-second keepalive comment
    is emitted when no event arrives to prevent proxy/browser timeout. The queue
    is cleaned up automatically when the student disconnects.
    """
    q: asyncio.Queue = asyncio.Queue()
    CLASS_STREAMS.setdefault(class_code, []).append(q)
    if class_code in CLASS_ASSIGNMENTS:
        await q.put({"type": "problem_assigned", "problem_id": CLASS_ASSIGNMENTS[class_code]})

    async def generator():
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    data = await asyncio.wait_for(q.get(), timeout=25)
                    yield f"data: {json.dumps(data)}\n\n"
                except asyncio.TimeoutError:
                    yield ": keepalive\n\n"
        finally:
            streams = CLASS_STREAMS.get(class_code, [])
            if q in streams:
                streams.remove(q)

    return StreamingResponse(
        generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/classroom/{class_code}/progress")
def classroom_progress(class_code: str):
    """
    Teacher endpoint: return each student's attempt summary for the live dashboard.

    Groups attempts by problem and correctness to produce correct/incorrect counts.
    Also includes the assigned level/difficulty if BKT has already run for that student.
    """
    db_path = ACTIVE_CLASSES.get(class_code)
    if db_path is None or not db_path.exists():
        return {"students": []}

    conn   = sqlite3.connect(db_path, timeout=10)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT session_id, student_id FROM sessions WHERE class_code = ?", (class_code,)
    )

    result = []
    for session_id, student_id in cursor.fetchall():
        cursor.execute("""
            SELECT problem_id, correctness, COUNT(*) as cnt
            FROM   steps
            WHERE  session_id = ?
            GROUP  BY problem_id, correctness
        """, (session_id,))
        problems: Dict[str, Dict[str, int]] = {}
        for problem_id, correctness, cnt in cursor.fetchall():
            if problem_id not in problems:
                problems[problem_id] = {"correct": 0, "incorrect": 0}
            if correctness == "CORRECT":
                problems[problem_id]["correct"] += cnt
            elif correctness == "INCORRECT":
                problems[problem_id]["incorrect"] += cnt

        total_correct   = sum(p["correct"]   for p in problems.values())
        total_incorrect = sum(p["incorrect"] for p in problems.values())

        assignment = cursor.execute(
            "SELECT level, difficulty FROM assignments WHERE session_id = ?", (session_id,)
        ).fetchone()

        try:
            hand_row = cursor.execute(
                "SELECT hand_raised FROM sessions WHERE session_id = ?", (session_id,)
            ).fetchone()
            hand_raised = bool(hand_row[0]) if hand_row else False
        except Exception:
            hand_raised = False

        knowledge_states = compute_knowledge_states(session_id, db_path)

        result.append({
            "student_id":       student_id,
            "session_id":       session_id,
            "problems":         problems,
            "total_correct":    total_correct,
            "total_incorrect":  total_incorrect,
            "assigned_to":      f"{assignment[0]} {assignment[1]}" if assignment else None,
            "hand_raised":      hand_raised,
            "knowledge_states": {kc: round(v, 3) for kc, v in knowledge_states.items()},
        })

    conn.close()
    return JSONResponse(
        content={"students": result},
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


def generate_results_report(class_code: str, db_path: Path) -> Optional[bytes]:
    """
    Build a plain Excel with one row per student showing, for each KC:
      initial P(L) (BKT over diagnostic exercises only),
      final P(L)   (BKT over all session exercises),
      delta        (final − initial),
      NLG          ((final − initial) / (1 − initial)).
    Returns raw bytes ready to stream, or None if no student has data yet.
    """
    try:
        import openpyxl
    except ImportError:
        print("[REPORT] openpyxl not installed — run: pip install openpyxl")
        raise

    kc_order = list(BKT_PARAMS.keys())
    kc_short  = {
        "move_constants":               "MC",
        "remove_coefficient":           "RC",
        "combine_like_terms":           "CLT",
        "expand_eliminate_parentheses": "EEP",
        "normalize_negative_sign":      "NNS",
    }

    diag_ids     = tuple(DIAGNOSTIC_EXERCISES)
    placeholders = ",".join("?" * len(diag_ids))

    conn     = sqlite3.connect(db_path, timeout=10)
    sessions = conn.execute(
        "SELECT session_id, student_id FROM sessions WHERE class_code = ?",
        (class_code,),
    ).fetchall()

    def _bkt(rows):
        states = {kc: BKT_PARAMS[kc]["p0"] for kc in BKT_PARAMS}
        for kc, correctness, hint_per_step, _ in rows:
            if kc not in BKT_PARAMS:
                continue
            correct  = (correctness == "CORRECT") and (not hint_per_step)
            states[kc] = bkt_update(states[kc], correct, BKT_PARAMS[kc])
        return {kc: round(v, 3) for kc, v in states.items()}

    report_rows = []
    for session_id, student_id in sessions:
        diag_steps = conn.execute(f"""
            SELECT kc, correctness, hint_per_step, timestamp
            FROM   steps
            WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
              AND  problem_id IN ({placeholders})
            ORDER  BY timestamp ASC
        """, (session_id, *diag_ids)).fetchall()

        all_steps = conn.execute("""
            SELECT kc, correctness, hint_per_step, timestamp
            FROM   steps
            WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
            ORDER  BY timestamp ASC
        """, (session_id,)).fetchall()

        if not all_steps:
            continue

        initial = _bkt(diag_steps)
        final   = _bkt(all_steps)

        row = {"student_id": student_id}
        for kc in kc_order:
            ini   = initial[kc]
            fin   = final[kc]
            delta = round(fin - ini, 3)
            nlg   = round((fin - ini) / (1.0 - ini), 3) if ini < 1.0 else 1.0
            row[f"inici_{kc}"]  = ini
            row[f"final_{kc}"]  = fin
            row[f"canvi_{kc}"]  = delta
            row[f"guany_{kc}"]  = nlg

        report_rows.append(row)

    conn.close()

    if not report_rows:
        return None

    kc_labels = {
        "move_constants":               "Move constants",
        "remove_coefficient":           "Remove coefficient",
        "combine_like_terms":           "Combine like terms",
        "expand_eliminate_parentheses": "Expand / eliminate parentheses",
        "normalize_negative_sign":      "Normalize negative sign",
    }

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary — one row per student ───────────────────────────────
    ws_all = wb.active
    ws_all.title = "Tots els alumnes"

    headers = ["Alumne"]
    for kc in kc_order:
        s = kc_short.get(kc, kc)
        headers += [f"{s} inici", f"{s} final", f"{s} canvi", f"{s} guany norm."]

    for col, h in enumerate(headers, 1):
        ws_all.cell(row=1, column=col, value=h)

    for ri, row in enumerate(report_rows, 2):
        ws_all.cell(row=ri, column=1, value=row["student_id"])
        col = 2
        for kc in kc_order:
            ws_all.cell(row=ri, column=col,     value=row[f"inici_{kc}"])
            ws_all.cell(row=ri, column=col + 1, value=row[f"final_{kc}"])
            ws_all.cell(row=ri, column=col + 2, value=row[f"canvi_{kc}"])
            ws_all.cell(row=ri, column=col + 3, value=row[f"guany_{kc}"])
            col += 4

    # ── Sheets 2…N: one per student ─────────────────────────────────────────
    for row in report_rows:
        stu = row["student_id"]
        # Excel sheet names are limited to 31 characters
        sheet_name = stu[:31]
        ws_stu = wb.create_sheet(title=sheet_name)

        # Header row
        stu_headers = ["Concepte", "Inici (diagnosi)", "Final (prova)", "Canvi", "Guany normalitzat"]
        for col, h in enumerate(stu_headers, 1):
            ws_stu.cell(row=1, column=col, value=h)

        # One row per KC
        for ri, kc in enumerate(kc_order, 2):
            ws_stu.cell(row=ri, column=1, value=kc_labels.get(kc, kc))
            ws_stu.cell(row=ri, column=2, value=row[f"inici_{kc}"])
            ws_stu.cell(row=ri, column=3, value=row[f"final_{kc}"])
            ws_stu.cell(row=ri, column=4, value=row[f"canvi_{kc}"])
            ws_stu.cell(row=ri, column=5, value=row[f"guany_{kc}"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.get("/api/classroom/{class_code}/report")
def download_report(class_code: str):
    """Generate and stream a per-student results Excel (initial/final P(L), delta, NLG)."""
    db_path = ACTIVE_CLASSES.get(class_code)
    if db_path is None:
        raise HTTPException(status_code=404, detail="Class not active.")

    try:
        data = generate_results_report(class_code, db_path)
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on the server.")

    if data is None:
        raise HTTPException(status_code=404, detail="Cap alumne té dades per generar el report.")

    ts       = datetime.now().strftime("%Hh%M_%d-%m-%Y")
    filename = f"resultats_{class_code}_{ts}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/classroom/{class_code}/end")
async def end_class(class_code: str):
    """
    Teacher endpoint: mark the session as ended and notify all connected students via SSE.

    Students poll GET /status; when they see ended=true they can start the final assessment.
    """
    ENDED_CLASSES.add(class_code)
    for q in CLASS_STREAMS.get(class_code, []):
        await q.put({"type": "session_ended"})
    print(f"[CLASS ENDED] {class_code}")
    return {"status": "ended", "class_code": class_code}


@app.post("/api/classroom/{class_code}/toggle-messages")
async def toggle_messages(class_code: str):
    """Teacher endpoint: toggle whether motivational messages are shown to students (A/B test)."""
    current = CLASS_MESSAGES_ENABLED.get(class_code, True)
    CLASS_MESSAGES_ENABLED[class_code] = not current
    return {"messages_enabled": CLASS_MESSAGES_ENABLED[class_code]}


@app.get("/api/classroom/{class_code}/status")
def class_status(class_code: str):
    """Students poll this to find out whether the teacher has ended the session."""
    return {
        "ended": class_code in ENDED_CLASSES,
        "messages_enabled": CLASS_MESSAGES_ENABLED.get(class_code, True),
    }


@app.get("/api/session/{session_id}/results")
def get_results(session_id: str):
    """
    Return per-KC knowledge states for the results page.

    initial_states: BKT run over only the 3 diagnostic exercises — reflects
                    the student's starting point before personalised practice.
    final_states:   BKT run over ALL attempts (diagnostic + personalised + final
                    assessment) — reflects the student's state at the end of the session.
    """
    db_path = get_db_for_session(session_id)
    if db_path is None or not db_path.exists():
        raise HTTPException(status_code=404, detail="Session not found.")

    diag_ids = tuple(DIAGNOSTIC_EXERCISES)
    placeholders = ",".join("?" * len(diag_ids))

    conn = sqlite3.connect(db_path, timeout=10)

    diag_rows = conn.execute(f"""
        SELECT kc, correctness, hint_per_step, timestamp
        FROM   steps
        WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
          AND  problem_id IN ({placeholders})
        ORDER  BY timestamp ASC
    """, (session_id, *diag_ids)).fetchall()

    all_rows = conn.execute("""
        SELECT kc, correctness, hint_per_step, timestamp
        FROM   steps
        WHERE  session_id = ? AND kc IS NOT NULL AND correctness IS NOT NULL
        ORDER  BY timestamp ASC
    """, (session_id,)).fetchall()

    conn.close()

    def _run(rows):
        states = {kc: p["p0"] for kc, p in BKT_PARAMS.items()}
        for kc, correctness, hint_per_step, _ in rows:
            if kc not in BKT_PARAMS:
                continue
            correct = (correctness == "CORRECT") and (not hint_per_step)
            states[kc] = bkt_update(states[kc], correct, BKT_PARAMS[kc])
        return {kc: round(v, 3) for kc, v in states.items()}

    initial = _run(diag_rows)
    final   = _run(all_rows)

    # Normalized Learning Gain: (post - pre) / (1 - pre)
    # Undefined when pre = 1.0 (no room to improve); clamped to 1.0 in that case.
    nlg = {
        kc: round((final[kc] - initial[kc]) / (1.0 - initial[kc]), 3)
        if initial[kc] < 1.0 else 1.0
        for kc in initial
    }

    return {
        "initial_states":           initial,
        "final_states":             final,
        "normalized_learning_gain": nlg,
    }


# -----------------------------------
# STATIC FILES
# -----------------------------------

DIST_DIR = PROJECT_ROOT / "frontend_react" / "dist"

# Serve CTAT exercise files (BRDs, HTML, CSS) under /CTAT.
if CTAT_DIR.exists():
    app.mount("/CTAT", StaticFiles(directory=CTAT_DIR), name="ctat")

if not DIST_DIR.exists():
    raise RuntimeError(f"Frontend build not found at {DIST_DIR}. Run: cd frontend_react && npm run build")

# Serve the compiled React SPA from /. html=True enables the SPA fallback so
# React Router handles client-side routes (/tutor, /teacher, /waiting).
app.mount("/", StaticFiles(directory=DIST_DIR, html=True), name="frontend")
